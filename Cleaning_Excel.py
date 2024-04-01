import openpyxl as xl
wb = xl.load_workbook('Covid_Deaths_New.xlsx')
Sheet = wb['Covid_Deaths']
for row in range(2, Sheet.max_row + 1):
    cell = Sheet.cell(row, 9)
    if cell.value is None:
        cell.value = 0


wb.save('Covid_Deaths_New.xlsx')